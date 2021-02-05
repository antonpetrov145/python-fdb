import openpyxl
import fdb

db = fdb.connect(
      host='host', database='base.fdb', user='user', password='password', charset='WIN1251'
)
cursor = db.cursor()

wb = openpyxl.load_workbook('file.xlsx')
ws = wb['sheet']

for i in range(2, ws.max_row+1)
  row = [cell.value for cell in ws[i]]
  cursor.execute("""INSERT INTO table (col1,col2,col3) VALUES (?,?,?)""", (str(row[0]), str(row[1]), str(row[2]), ))

db.commit()
db.cloes()
